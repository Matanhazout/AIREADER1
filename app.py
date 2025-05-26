import os
import csv
import base64
import imghdr
import re
import json
from flask import Flask, render_template, request, jsonify, redirect, url_for, make_response
from PyPDF2 import PdfReader
import docx
import openpyxl
import chardet
from difflib import SequenceMatcher

app = Flask(__name__)

DATA_DIRS = {
    'guest': 'data_guest',
    'admin': 'data_admin'
}

USERS_FILE = 'users.json'
USERS = {}

# יצירת תיקיות אם לא קיימות
for d in DATA_DIRS.values():
    if not os.path.exists(d):
        os.makedirs(d)

# טעינת משתמשים מקובץ JSON
def load_users():
    global USERS
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            USERS = json.load(f)
    else:
        # יצירת משתמשים ראשוניים אם אין קובץ
        USERS.update({
            'guest': {'password': '12345', 'role': 'guest'},
            'shay': {'password': '12345', 'role': 'guest'},
            'udi': {'password': '12345', 'role': 'guest'},
            'matan': {'password': '12345', 'role': 'guest'},
            'yehuda': {'password': '12345', 'role': 'guest'},
            'harel': {'password': '12345', 'role': 'guest'},
            'admin': {'password': '12345', 'role': 'admin'}
        })
        save_users()

def save_users():
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(USERS, f, ensure_ascii=False, indent=2)

load_users()


# --- פונקציות קריאת קבצים כפי שנתת במקור ---

def read_docx(file_path):
    if os.path.basename(file_path).startswith('~$'):
        return "", []

    doc = docx.Document(file_path)
    content = ''
    images = []

    for para in doc.paragraphs:
        content += para.text + '\n'

    for rel in doc.part.rels.values():
        if "image" in rel.target_ref:
            image_data = rel.target_part.blob
            image_type = imghdr.what(None, h=image_data) or 'jpeg'
            img_base64 = base64.b64encode(image_data).decode('utf-8')
            images.append(f"data:image/{image_type};base64,{img_base64}")

    return content, images


def read_csv(file_path):
    content = ''
    with open(file_path, 'rb') as file:
        raw_data = file.read(10000)
        result = chardet.detect(raw_data)
        encoding = result['encoding']

    try:
        with open(file_path, 'r', encoding=encoding) as file:
            reader = csv.reader(file)
            for row in reader:
                content += ' | '.join(row) + '\n'
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='ISO-8859-1') as file:
            reader = csv.reader(file)
            for row in reader:
                content += ' | '.join(row) + '\n'

    return content


def read_pdf(file_path):
    with open(file_path, 'rb') as file:
        reader = PdfReader(file)
        content = ''
        for page in reader.pages:
            content += page.extract_text() or ''
        return content


def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    content = ''
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                content += str(cell.value or '') + ' '
    return content


def preprocess_question(question):
    question = question.lower()
    question = re.sub(r'[^\w\s]', '', question)
    question = re.sub(r'\s+', ' ', question).strip()
    return question


def get_words(text):
    return re.findall(r'\b\w+\b', text)


def keyword_similar_or_surrounded(keyword, word, threshold=0.9):
    keyword = keyword.lower()
    word = word.lower()

    # השוואת מילים שקולות ל־IP
    ip_variants = {'ip', 'הip', 'ה-ip', 'ip.', 'ip?', 'ip!', 'הip.', 'הip?', 'הip!'}
    if (keyword in ip_variants and word in ['כתובת', 'כתובת ip', 'אייפי']) or \
       (word in ip_variants and keyword in ['כתובת', 'כתובת ip', 'אייפי']):
        return True

    # תנאים מיוחדים
    if ((keyword in ['מי', 'מה', 'תן'] and word == 'מידע')
            or (word in ['מי', 'מה', 'תן'] and keyword == 'מידע')):
        return True

    if len(keyword) < 3 or len(word) < 3:
        return False

    if keyword in word or word in keyword:
        return True

    if re.fullmatch(rf'.{{0,1}}{re.escape(keyword)}.{{0,1}}', word):
        return True

    return SequenceMatcher(None, keyword, word).ratio() >= threshold

def find_best_match(question, data_dir):
    files = os.listdir(data_dir)
    question_clean = preprocess_question(question)
    keywords = question_clean.split()

    best_file = None
    best_filename = ""
    highest_match_count = 0

    for filename in files:
        full_path = os.path.join(data_dir, filename)
        if os.path.isfile(full_path):
            filename_base = preprocess_question(os.path.splitext(filename)[0])
            filename_words = set(filename_base.split())

            match_count = 0
            for keyword in keywords:
                for word in filename_words:
                    if keyword_similar_or_surrounded(keyword, word):
                        match_count += 1

            if match_count > highest_match_count:
                highest_match_count = match_count
                best_file = full_path
                best_filename = filename

    if not best_file:
        return None

    content = ''
    images = []
    if best_file.endswith('.txt'):
        with open(best_file, 'r', encoding='utf-8') as f:
            content = f.read()
    elif best_file.endswith('.csv'):
        content = read_csv(best_file)
    elif best_file.endswith('.pdf'):
        content = read_pdf(best_file)
    elif best_file.endswith('.docx'):
        content, images = read_docx(best_file)
    elif best_file.endswith('.xlsx'):
        content = read_excel(best_file)

    if best_file.endswith('.docx'):
        return {
            'filename': best_filename,
            'content': content,
            'images': images
        }

    sections = re.findall(r'<section.*?>(.*?)</section>', content,
                          re.DOTALL | re.IGNORECASE)
    matched_sections = []
    question_keywords = set(keywords)

    for section in sections:
        section_clean = preprocess_question(section)

        h1_match = False
        h1_matches = re.findall(r'<h1.*?>(.*?)</h1>', section,
                                re.DOTALL | re.IGNORECASE)
        for h1 in h1_matches:
            h1_clean = preprocess_question(h1)
            h1_words = get_words(h1_clean)
            for keyword in question_keywords:
                if any(keyword_similar_or_surrounded(keyword, w) for w in h1_words):
                    h1_match = True
                    break
            if h1_match:
                break

        comment_match = False
        comments = re.findall(r'<!--(.*?)-->', section, re.DOTALL)
        for comment in comments:
            comment_clean = preprocess_question(comment)
            comment_words = get_words(comment_clean)
            for keyword in question_keywords:
                if any(keyword_similar_or_surrounded(keyword, w) for w in comment_words):
                    comment_match = True
                    break
            if comment_match:
                break

        strong_lines = []
        lines = section.splitlines()
        for line in lines:
            strong_words = re.findall(r'<strong>(.*?)</strong>', line, re.IGNORECASE)
            strong_words_clean = []
            for s in strong_words:
                s_clean = preprocess_question(s)
                strong_words_clean.extend(get_words(s_clean))
            if any(keyword_similar_or_surrounded(keyword, word) for keyword in question_keywords for word in strong_words_clean):
                strong_lines.append(line.strip())

        if (h1_match or comment_match) and strong_lines:
            matched_sections.extend(strong_lines)
        elif h1_match or comment_match:
            matched_sections.append(section.strip())

    return_all = any(word in {'כל', 'הכל', 'כולם'} for word in question_clean.split())

    if matched_sections:
        if return_all:
            final_content = '\n\n'.join(matched_sections)
        else:
            final_content = matched_sections[0]
    else:
        final_content = 'אין מידע על כך.'

    return {
        'filename': best_filename,
        'content': final_content,
        'images': images
    }


# --- ראוטים ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if username in USERS and USERS[username]['password'] == password:
            response = make_response(redirect(url_for('index')))
            response.set_cookie('username', username)
            return response
        else:
            error = "שם משתמש או סיסמה שגויים"
            return render_template('login.html', error=error)

    return render_template('login.html')


@app.route('/logout')
def logout():
    response = make_response(redirect(url_for('login')))
    response.delete_cookie('username')
    return response


@app.route('/')
def index():
    username = request.cookies.get('username')
    if not username or username not in USERS:
        return redirect(url_for('login'))
    role = USERS[username]['role']
    return render_template('index.html', username=username, role=role)


@app.route('/ask', methods=['POST'])
def ask():
    username = request.cookies.get('username')
    if not username or username not in USERS:
        return jsonify({'answer': 'נא להתחבר.', 'images': []})

    role = USERS[username]['role']
    data_dir = DATA_DIRS.get(role, 'data_guest')

    question = request.form.get('question')

    best_match = find_best_match(question, data_dir)

    if best_match:
        return jsonify({
            'answer': best_match['content'],
            'filename': best_match['filename'],
            'images': best_match.get('images', [])
        })
    else:
        return jsonify({'answer': 'אין תשובה מתאימה.', 'images': []})


# ראוט להוספת משתמש חדש (רק למנהלים)
@app.route('/admin/add_user', methods=['GET', 'POST'])
def add_user():
    username = request.cookies.get('username')
    if not username or username not in USERS or USERS[username]['role'] != 'admin':
        return jsonify({'error': 'אין הרשאה'}), 403

    if request.method == 'POST':
        new_username = request.form.get('new_username', '').strip()
        new_password = request.form.get('new_password', '').strip()
        new_role = request.form.get('new_role', 'guest').strip()

        if not new_username or not new_password:
            return jsonify({'error': 'יש למלא שם משתמש וסיסמה'}), 400

        if new_username in USERS:
            return jsonify({'error': 'שם משתמש כבר קיים'}), 400

        if new_role not in ['guest', 'admin']:
            return jsonify({'error': 'תפקיד לא תקין'}), 400

        USERS[new_username] = {'password': new_password, 'role': new_role}
        save_users()

        return jsonify({'success': 'משתמש נוסף בהצלחה'})

    # GET - הצגת טופס בסיסי להוספת משתמש (אפשר לשפר לתבנית HTML נפרדת)
    return '''
    <h2>הוספת משתמש חדש</h2>
    <form method="post">
      שם משתמש: <input name="new_username"><br>
      סיסמה: <input type="password" name="new_password"><br>
      תפקיד: <select name="new_role">
        <option value="guest">Guest</option>
        <option value="admin">Admin</option>
      </select><br>
      <input type="submit" value="הוסף משתמש">
    </form>
    '''


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000, debug=True)